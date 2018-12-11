#!/usr/bin/env python3

import openpyxl
import sys
import render_pdf
import xlrd

def unwrap_strengths(unw):
    lines = unw.split('\n')
    return [u.split('.')[1].strip() for u in lines]

def BOA_and_McCollum(ws, output_dir):
    for row in ws.iter_rows(min_row=3):
        mentor_name = row[0].value
        mentor_strengths = unwrap_strengths(row[1].value)
        mentee_name = row[3].value
        mentee_strengths = unwrap_strengths(row[4].value)
        fname = '{}/{}.pdf'.format(output_dir, mentor_name)
        render_pdf.create_name_tent(fname, mentor_name, mentor_strengths)
        fname = '{}/{}.pdf'.format(output_dir, mentee_name)
        render_pdf.create_name_tent(fname, mentee_name, mentee_strengths)

def McCollum_HEB_Student_Strengths(ws, output_dir):
    for row in ws.iter_rows(min_row=2):
        mentee_name = row[0].value
        if mentee_name is None:
            continue
        mentee_strengths = [row[1].value, row[2].value, row[3].value]
        fname = '{}/{}.pdf'.format(output_dir, mentee_name)
        render_pdf.create_name_tent(fname, mentee_name, mentee_strengths)

def InspireU_Mentor_Strengths_HEB_McCollum(ws, output_dir):
    for row in ws.iter_rows(min_row=2):
        name = row[0].value
        if name is None:
            continue
        strengths = [row[n].value for n in range(1,6)]
        if None in strengths:
            continue
        fname = '{}/{}.pdf'.format(output_dir, name)
        render_pdf.create_name_tent(fname, name, strengths)

def HEB_Remaining(ws, output_dir):
    for row in ws.iter_rows():
        name = row[0].value
        if name is None:
            continue
        if row[4].value is None:
            upper = 4
        else:
            upper = 6
        strengths = [row[n].value for n in range(1,upper)]
        if None in strengths:
            continue
        fname = '{}/{}.pdf'.format(output_dir, name)
        render_pdf.create_name_tent(fname, name, strengths)

def Harlandale_HEB_Student_Strengths_Tracker(ws, output_dir):
    for row in ws.iter_rows():
        name = row[0].value
        if name is None:
            continue
        strengths = [row[n].value for n in range(1,4)]
        if None in strengths:
            continue
        fname = '{}/{}.pdf'.format(output_dir, name)
        render_pdf.create_name_tent(fname, name, strengths)

def InspireU_Mentor_Strengths_HEB_Harlandale_HS(ws, output_dir):
    for row in ws.iter_rows():
        name = row[0].value
        if name is None:
            continue
        strengths = [row[n].value for n in range(1,6)]
        if None in strengths:
            continue
        fname = '{}/{}.pdf'.format(output_dir, name)
        render_pdf.create_name_tent(fname, name, strengths)

def generic(ws, output_dir):
    # try the age-old one line per person without header line
    # name in the first column
    # strengths in the next columns - 3 for some and 5 for others
    for row in ws.iter_rows():
        name = row[0].value
        if name is None:
            continue
        strengths = [row[n].value for n in range(1,ws.max_column) if row[n].value is not None]
        fname = '{}/{}.pdf'.format(output_dir, name)
        render_pdf.create_name_tent(fname, name, strengths)

def full34(ws, output_dir, image=None):
    # try the Gallup format: one line per person with header line
    # name in the first column
    # strengths in the next columns - all 34 are avail, but only print 10
    #TODO: add in 5 or 10 option
    first_row = True
    for row in ws.get_rows():
        if first_row:
            first_row = False
            continue
        name = row[0].value
        if name is '':
            return
        # name is last, first
        try:
            name_split = name.split(', ')
            last_name = ', '.join(name_split[:-1])
            first_name = name_split[-1]
        except ValueError:
            print("Error splitting full name: {}".format(full_name))
            continue
        name = '{} {}'.format(first_name, last_name)
        strengths = [row[n].value for n in range(1,11) if row[n].value is not None]
        fname = '{}/{}.pdf'.format(output_dir, name)
        render_pdf.create_name_tent(fname, name, strengths, image=image)

def main(xl_fname, output_dir, proc=None, image=None):
    extension = xl_fname.split('.')[-1]
    if extension == 'xlsx':
        wb = openpyxl.load_workbook(xl_fname, read_only=True)
        ws = wb.active
    elif extension == 'xls':
        wb = xlrd.open_workbook(xl_fname)
        ws = wb.sheet_by_index(0)
    if proc == 'BOA and McCollum':
        BOA_and_McCollum(ws, output_dir)
    elif proc == 'McCollum HEB Student Strengths':
        McCollum_HEB_Student_Strengths(ws, output_dir)
    elif proc == 'InspireU Mentor Strengths- HEB McCollum':
        InspireU_Mentor_Strengths_HEB_McCollum(ws, output_dir)
    elif proc == 'HEB-remaining':
        HEB_Remaining(ws, output_dir)
    elif proc == 'Harlandale HEB Student Strengths Tracker':
        Harlandale_HEB_Student_Strengths_Tracker(ws, output_dir)
    elif proc == 'InspireU Mentor Strengths- HEB Harlandale HS':
        InspireU_Mentor_Strengths_HEB_Harlandale_HS(ws, output_dir)
    elif proc == 'Holmes Acelity Student Strengths Tracker':
        Harlandale_HEB_Student_Strengths_Tracker(ws, output_dir)
    else:
        full34(ws, output_dir, image)

def alp():
    with open(sys.argv[1], 'r') as infile:
        names = infile.read().splitlines()
    output_dir = sys.argv[2]
    image = sys.argv[3]
    for name in names:
        fname = '{}/{}.pdf'.format(output_dir, name)
        render_pdf.create_name_tent(fname, name, [], image=image)

if __name__ == '__main__':
    # args: xlsx pdf_dir [png]
    alp()
    raise SystemExit()
    proc = sys.argv[1].split('/')[-1].split('.')[0]
    if len(sys.argv) == 4:
        image = sys.argv[3]
    else:
        image = None
    main(sys.argv[1], sys.argv[2], proc, image)
